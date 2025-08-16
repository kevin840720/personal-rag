# -*- encoding: utf-8 -*-
"""
@File    :  office_excel.py
@Time    :  2025/06/18 16:15:40
@Author  :  Kevin Wang
@Desc    :  用 Docling 處理 Excel，產出 LoaderResult
"""

from io import BytesIO
from pathlib import Path
from typing import (Any,
                    List,
                    Union,
                    )


from ingestion.base import DocumentLoader, LoaderResult
from ingestion.utils import export_to_user_text
from docling.backend.abstract_backend import (DeclarativeDocumentBackend,
                                              PaginatedDocumentBackend,
                                              )
from docling.backend.msexcel_backend import MsExcelDocumentBackend
from docling.backend.msexcel_backend import ExcelTable
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import InputDocument
from docling.document_converter import (DocumentConverter,
                                        FormatOption,
                                        )
from docling.pipeline.simple_pipeline import SimplePipeline
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing_extensions import override

from objects import (DocumentMetadata,
                     FileType,
                     )
from pathlib import Path
from typing import List, Union
from openpyxl import load_workbook

from ingestion.base import DocumentLoader, LoaderResult
from ingestion.utils import export_to_user_text
from objects import DocumentMetadata, FileType

from docling.backend.abstract_backend import (DeclarativeDocumentBackend,
                                              PaginatedDocumentBackend,
                                              )
from docling.document_converter import DocumentConverter, FormatOption
from docling.datamodel.base_models import InputFormat
from docling.pipeline.simple_pipeline import SimplePipeline
from docling.backend.msexcel_backend import MsExcelDocumentBackend


class MsExcelDocumentDataOnlyBackend(MsExcelDocumentBackend):
    """由於 Docling 內建的 MS Excel Backend 會讀取「公式」而非「值」，所以另外新增一個讀取「值」的版本 

    Args:
        MsExcelDocumentBackend (_type_): _description_

    Raises:
        RuntimeError: _description_
    """
    @override
    def __init__(self,
                 in_doc:"InputDocument",
                 path_or_stream:Union[BytesIO,Path],
                 ) -> None:
        DeclarativeDocumentBackend.__init__(self, in_doc, path_or_stream)
        PaginatedDocumentBackend.__init__(self, in_doc, path_or_stream)

        # 以下是從 docling.backend.msexcel_backend.MsExcelDocumentBackend 搬來的
        # --------------------------------------------------------------------------------------------------------------
        # max_levels / parents 應該是為「階層結構」(Hierarchy) 的預留空間，目前只用到第 0 層，目前沒啥用的樣子
        self.max_levels = 10
        self.parents: dict[int, Any] = {}
        for i in range(-1, self.max_levels):
            self.parents[i] = None

        self.workbook = None
        try:
            if isinstance(self.path_or_stream, BytesIO):
                self.workbook = load_workbook(filename=self.path_or_stream,
                                              data_only=True,  # True 的話會讀取「值」而非「公式」
                                              )

            elif isinstance(self.path_or_stream, Path):
                self.workbook = load_workbook(filename=str(self.path_or_stream),
                                              data_only=True,  # True 的話會讀取「值」而非「公式」
                                              )

            self.valid = self.workbook is not None
        except Exception as err:
            self.valid = False
            raise RuntimeError(f"MsExcelDocumentBackend could not load document with hash {self.document_hash}") from err
        # --------------------------------------------------------------------------------------------------------------
        
        # 每個 sheet 讀取的 row/col 上限
        self.max_rows = 10000
        self.max_cols = 10000

    @override
    def _find_data_tables(self, sheet:Worksheet) -> list[ExcelTable]:
        """尋找「連續資料儲存格」，並將其視作一個 Table

        Args:
            sheet: The Excel worksheet to be parsed.

        Returns:
            A list of ExcelTable objects representing the data tables.
        """
        tables: list[ExcelTable] = []  # List to store found tables
        visited: set[tuple[int, int]] = set()  # Track already visited cells

        # Iterate over all cells in the sheet
        for ri, row in enumerate(sheet.iter_rows(values_only=False)):

            # 新增最大讀取 row/column 數量
            if ri >= self.max_rows:
                break
            for rj, cell in enumerate(row):
                if rj >= self.max_cols:
                    break

                # Skip empty or already visited cells
                if cell.value is None or (ri, rj) in visited:
                    continue

                # If the cell starts a new table, find its bounds
                table_bounds, visited_cells = self._find_table_bounds(sheet, ri, rj)

                visited.update(visited_cells)  # Mark these cells as visited
                tables.append(table_bounds)

        return tables

class DoclingExcelLoader(DocumentLoader):
    """用 Docling 解析 Excel，產出 LoaderResult"""

    def __init__(self, data_only: bool = True):
        super().__init__()
        if data_only:
            self.converter = DocumentConverter(allowed_formats=[InputFormat.XLSX],
                                               format_options={InputFormat.XLSX: FormatOption(pipeline_cls=SimplePipeline,
                                                                                              backend=MsExcelDocumentDataOnlyBackend,
                                                                                              )
                                                               }
                                               )
        else:
            self.converter = DocumentConverter()

    def _get_metadata(self,
                      path:Union[Path,str],
                      ) -> DocumentMetadata:
        wb = load_workbook(str(path), read_only=True)
        props = wb.properties
        return DocumentMetadata(file_type=FileType.XLSX,
                                file_name=Path(path).name,
                                title=props.title,
                                author=props.creator,
                                subject=props.subject,
                                created_at=props.created,
                                modified_at=props.modified,
                                )

    def load(self,
             path:Union[str,Path],
             ) -> List[LoaderResult]:
        path = Path(path)
        docling_doc = self.converter.convert(str(path)).document
        metadata = self._get_metadata(path)
        content = export_to_user_text(docling_doc)
        result = LoaderResult(content=content,
                              metadata=metadata,
                              doc=docling_doc,
                              )
        return [result]
