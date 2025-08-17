# -*- encoding: utf-8 -*-
"""
@File    :  test_office_excel.py
@Time    :  2025/08/16 20:29:34
@Author  :  Kevin Wang
@Desc    :  None
"""

# -*- encoding: utf-8 -*-
"""
@File    :  test_office_excel.py
@Time    :  2025/08/16 18:10:00
@Author  :  Kevin Wang
@Desc    :  Tests for DoclingExcelLoader (returns LoaderResult)
"""

from datetime import datetime, timezone
from pathlib import Path
from typing import List

import pytest
from openpyxl import Workbook  # docling 的 dependence
from docling_core.types.doc.document import DoclingDocument

from conftest import SKIP_REAL_FILE_TEST
from ingestion.base import LoaderResult
from ingestion.file_loaders.office_excel import DoclingExcelLoader
from objects import FileType


@pytest.fixture
def test_xlsx_path(tmp_path: Path) -> Path:
    """建立一個臨時的 XLSX 檔案，供單一檔案測試使用"""
    file_path = tmp_path / "test01.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "測試工作表"
    ws["A1"] = "品名"
    ws["B1"] = "數量"
    ws["C1"] = "單價"
    ws["D1"] = "小計"
    ws["A2"] = "蘋果"
    ws["B2"] = 3
    ws["C2"] = 10
    ws["D2"] = 30

    # 設定文件屬性（metadata）
    props = wb.properties
    now = datetime.now(timezone.utc)
    props.title = "單檔測試標題"
    props.creator = "Kevin Wang"
    props.subject = "單檔測試主題"
    props.created = now
    props.modified = now

    wb.save(str(file_path))
    return file_path


@pytest.fixture
def test_xlsx_paths(tmp_path: Path) -> List[Path]:
    """建立多個臨時的 XLSX 檔案（pytest 測試結束後會自動刪除）"""
    paths: List[Path] = []
    now = datetime.now(timezone.utc)

    # file1
    p1 = tmp_path / "test01.xlsx"
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "測試表一"
    ws1["A1"] = "商品"
    ws1["A2"] = "A"
    props1 = wb1.properties
    props1.title = "批次檔案一"
    props1.creator = "Kevin Wang"
    props1.subject = "批次測試"
    props1.created = now
    props1.modified = now
    wb1.save(str(p1))
    paths.append(p1)

    # file2
    p2 = tmp_path / "test02.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "測試表二"
    ws2["A1"] = "商品"
    ws2["A2"] = "B"
    props2 = wb2.properties
    props2.title = "批次檔案二"
    props2.creator = "Kevin Wang"
    props2.subject = "批次測試"
    props2.created = now
    props2.modified = now
    wb2.save(str(p2))
    paths.append(p2)

    return paths


class TestDoclingExcelLoader:
    @pytest.fixture
    def xlsx_loader(self):
        # 預設 data_only=True，符合 loader 設計（讀取儲存格值）
        return DoclingExcelLoader()

    def test_load(self,
                  xlsx_loader,
                  test_xlsx_path:Path,
                  ):
        results = xlsx_loader.load(test_xlsx_path)
        assert isinstance(results, list) and len(results) == 1

        res = results[0]
        assert isinstance(res, LoaderResult)

        # 檢驗 metadata
        assert res.metadata.file_type == FileType.XLSX
        assert res.metadata.file_name == test_xlsx_path.name
        assert isinstance(res.metadata.created_at, datetime)
        assert isinstance(res.metadata.modified_at, datetime)

        # content 與 doc 基本檢查（包含原文關鍵字）
        assert isinstance(res.content, str)
        assert "品名" in res.content or "蘋果" in res.content
        assert isinstance(res.doc, DoclingDocument)

    def test_load_multi(self,
                        xlsx_loader,
                        test_xlsx_paths: List[Path]):
        results = xlsx_loader.load_multi(test_xlsx_paths)

        assert isinstance(results, list)
        assert len(results) == len(test_xlsx_paths)

        for res, src_path in zip(results, test_xlsx_paths):
            assert isinstance(res, LoaderResult)

            # metadata 檢查
            assert res.metadata.file_type == FileType.XLSX
            assert res.metadata.file_name == src_path.name
            assert isinstance(res.metadata.created_at, datetime)
            assert isinstance(res.metadata.modified_at, datetime)

            # content 與 doc 基本檢查
            assert isinstance(res.content, str)
            assert "商品" in res.content or "測試" in res.content
            assert isinstance(res.doc, DoclingDocument)

    @pytest.mark.skipif(SKIP_REAL_FILE_TEST, reason="Skipping real file tests")
    def test_load_real(self, xlsx_loader):
        """檢驗讀取實際 xlsx 檔案時，有沒有報錯
        使用不同檔案時，可能需要調整一下測試內容
        """
        file_path = Path("./tests/data/RAG測試文件.xlsx")
        results = xlsx_loader.load(file_path)
        assert isinstance(results, list) and len(results) == 1

        res = results[0]
        assert isinstance(res, LoaderResult)

        # 檢驗 metadata
        assert res.metadata.file_type == FileType.XLSX
        assert res.metadata.file_name == file_path.name
        assert isinstance(res.metadata.created_at, datetime)
        assert isinstance(res.metadata.modified_at, datetime)

        # content 與 doc 基本檢查（包含原文關鍵字）
        assert isinstance(res.content, str) and ("實際支付金額" in res.content)
        assert isinstance(res.doc, DoclingDocument)
